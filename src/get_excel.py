import xlrd
import aiohttp
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import xlrd
import  asyncio

def is_xlsx(content: bytes) -> bool:
    """判断是否是 xlsx 文件（zip 文件，以 PK 开头）"""
    return content[:2] == b'PK'

async def get_excel_data(excel_url, timeout=60):
    try:
        # 下载 Excel 内容
        async with aiohttp.ClientSession() as session:
            async with session.get(excel_url, timeout=timeout, headers={'User-Agent': 'Mozilla/5.0'}) as response:
                response.raise_for_status()
                content = await response.read()

        excel_buffer = BytesIO(content)

        if is_xlsx(content):
            # ✅ 处理 .xlsx，忽略隐藏 sheet
            wb = load_workbook(excel_buffer, read_only=True)
            visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
            wb.close()

            excel_buffer.seek(0)
            excel_data = pd.read_excel(excel_buffer, sheet_name=visible_sheets, header=None)

            result = [(sheet_name, df) for sheet_name, df in excel_data.items()]
            return True, result

        else:
            # ✅ 处理 .xls，判断 sheet.visibility
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
            visible_sheet_names = [
                sheet.name for sheet in workbook.sheets()
                if getattr(sheet, 'visibility', 0) == 0  # 0 表示可见
            ]
            excel_data = pd.read_excel(BytesIO(content), sheet_name=visible_sheet_names, header=None)

            result = [(sheet_name, df) for sheet_name, df in excel_data.items()]
            return True, result

    except Exception as e:
        return False, str(e)

#
# def is_xlsx(content: bytes) -> bool:
#     """判断是否是 xlsx 文件（zip 文件，以 PK 开头）"""
#     return content[:2] == b'PK'
#
# async def get_excel_data(excel_url, timeout=60):
#     try:
#         # 下载 Excel 内容
#         async with aiohttp.ClientSession() as session:
#             async with session.get(excel_url, timeout=timeout, headers={'User-Agent': 'Mozilla/5.0'}) as response:
#                 response.raise_for_status()
#                 content = await response.read()
#
#         excel_buffer = BytesIO(content)
#
#         if is_xlsx(content):
#             # ✅ 处理 .xlsx，忽略隐藏 sheet
#             wb = load_workbook(excel_buffer, read_only=True)
#             visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
#             wb.close()
#
#             excel_buffer.seek(0)
#             excel_data = pd.read_excel(excel_buffer, sheet_name=visible_sheets, header=None, engine='openpyxl')
#             return True, [df for df in excel_data.values()]
#         else:
#             # ✅ 处理 .xls，判断 sheet.visibility
#             workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
#             visible_sheet_names = [
#                 sheet.name for sheet in workbook.sheets()
#                 if getattr(sheet, 'visibility', 0) == 0  # 0 表示可见
#             ]
#             excel_data = pd.read_excel(BytesIO(content), sheet_name=visible_sheet_names, header=None, engine='xlrd')
#             return True, [df for df in excel_data.values()]
#     except Exception as e:
#         return False, str(e)
# excel_url3 = r'https://csde-file.trinablue.com/prod/rsbu-mdm-contract-ps-api/20250709/75261ABAA2B64E43B634CB870F181B83_686e37d4e4b0b02ffdfbb8be.xlsx'
#
# x = asyncio.run(get_excel_data(excel_url3))
# print(x)