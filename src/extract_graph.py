import json
import time
import asyncio
from typing import Literal
from sys import exception
from typing import Annotated
from pydantic import BaseModel, Field
from src.utils import count_token, extract_json,get_window,get_type_two_window,get_parse_window,extract_json_list
from langgraph.graph import StateGraph, MessagesState, START, END
from typing import TypedDict, Any
from langgraph.types import Command
from src.llm_singleton import OpenAIClientSingleton
from src.token_singleton import TokenizerSingleton
from src.config import Config
import pandas as pd
import aiofiles
import openpyxl
from modelscope import AutoModelForCausalLM, AutoTokenizer

class State(TypedDict):
    type_window:int
    llm_window:int

    excel_df:Any
    excel_dict:Any
    excel_type: dict | None
    flag :str

    type_one_head:dict | None
    type_one_index:dict | None
    type_one_common:dict | None

    type_two_exe_prompts :Any
    chunks:Any
    result:list | None

    error:str


class PreProcessNode:
    def __call__(self, state: State):
        df = state["excel_df"]
        df.dropna(how="all", inplace=True)
        df.dropna(axis=1, how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)
        excel_dict = json.loads(df.to_json(force_ascii=False, orient="index"))
        clean_dict = {}
        for key, value in excel_dict.items():
            d_clean = {k: v for k, v in value.items() if v is not None}
            clean_dict[key] = d_clean
        excel_list = list(excel_dict.items())
        return Command(update={"excel_dict": clean_dict,"excel_list":excel_list}, goto="FineTypeBot")

class FineTypeNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    async def __call__(self, state: State):
        llm_window = state["llm_window"]
        type_window =  state["type_window"]
        excel_dict = state["excel_dict"]
        user_prompt = f"Json格式的Excel表格输入:"
        flag,message = get_window(excel_dict,type_window,self.system_prompt,user_prompt)

        try:
            response =await OpenAIClientSingleton().create_completion(messages=message)

            result = response.choices[0].message.content
            token_num = response.usage.prompt_tokens
            result_json = extract_json(result)

            if result_json["type"] == 1 :
                return Command(update={"excel_type": result_json}, goto="TypeOneGetHeadBot")
            elif result_json["type"] == 2 and token_num>=llm_window:
                return Command(update={"excel_type": result_json,"flag":"通过切片"}, goto="TypeTwoGetChunksBot")
            elif result_json["type"] == 2 and token_num<llm_window:
                return Command(update={"excel_type": result_json,"flag":"直接送入模型"}, goto="OnlyllmBot")
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"FineTypeNode出错：{exception_message}"}, goto=END)
class TypeOneGetHeadNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    async def __call__(self, state: State):
        llm_window = state["llm_window"]
        head_index = 30
        excel_dict=dict(list(state["excel_dict"].items())[:head_index])
        user_prompt = "json格式的excel文件:"
        flag, message = get_window(excel_dict, llm_window, self.system_prompt, user_prompt)
        try:
            response =await OpenAIClientSingleton().create_completion(messages=message)
            result = response.choices[0].message.content
            result_json = extract_json(result)
            return Command(update={"type_one_head": result_json}, goto="TypeOneGetIndexBot")
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"TypeOneGetHeadNode出错：{exception_message}"}, goto=END)
class TypeOneGetIndexNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    async def __call__(self, state: State):
        llm_window = state["llm_window"]
        head_index = 30
        excel_dict = dict(list(state["excel_dict"].items())[:head_index])
        user_prompt = "json格式的excel文件:"
        flag, message = get_window(excel_dict, llm_window, self.system_prompt, user_prompt)
        try:
            response = await OpenAIClientSingleton().create_completion(messages=message)
            result = response.choices[0].message.content
            result_json = extract_json(result)
            return Command(update={"type_one_index": result_json}, goto="TypeOneGetCommonBot")
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"TypeOneGetIndexNode出错：{exception_message}"}, goto=END)
class TypeOneGetCommonNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    async def __call__(self, state: State):
        llm_window = state["llm_window"]
        head_index=state["type_one_head"]["index"]
        excel_dict = dict(list(state["excel_dict"].items())[:head_index+1])
        user_prompt = f"json格式的excel表头:{excel_dict}"
        message = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": user_prompt}]
        try:
            response = await OpenAIClientSingleton().create_completion(messages=message)
            result = response.choices[0].message.content
            result_json = extract_json(result)
            return Command(update={"type_one_common": result_json}, goto="TypeOneExeBot")
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"TypeOneGetCommonNode出错：{exception_message}"}, goto=END)
class TypeOneExeNode:
     async def __call__(self, state: State):
         # async with aiofiles.open(output_filename, 'w', encoding='utf-8') as json_file:
         #     # 使用异步json.dumps
         #     await json_file.write(jsonsr)
         count=0
         result=[]
         index = state["type_one_index"]
         extra = state["type_one_common"]
         start_row = state["type_one_head"]["index"]
         excel_list = list(state["excel_dict"].values())[start_row+1:]
         try:
             for i,data in enumerate(excel_list):
                 temp = {
                     key: data.get(str(value)) if value is not None else None
                     for key, value in index.items()
                 }
                 temp = self.merge_dicts(temp,extra)
                 if temp["generation_account"]!=None or temp["transaction_id"]!=None:
                    if i ==0:
                        count=self.count_non_empty_fields(temp)
                        result.append(temp)
                    else:
                        if self.count_non_empty_fields(temp)>=count:
                            result.append(temp)
                        else:
                            continue

             # with open(output_filename, 'w', encoding='utf-8') as json_file:
             #     json.dump(result, json_file, ensure_ascii=False, indent=4)
             return Command(update={"result":result},goto=END)
         except Exception as e:
                exception_message = str(e)
                return Command(update={"error": f"TypeOneExeNode出错：{exception_message}"}, goto=END)

     def merge_dicts(self,dict1,dict2):  #
         for key, value in dict2.items():  # 遍历第二个字典
             if key in dict1:  # 如果第一个字典已有该键
                 if value!=None:  # 如果第一个字典的值是None
                     dict1[key] = value  # 用第二个字典的值替换
             else:
                 dict1[key] = value  # 如果第一个字典没有该键，直接添加
         return dict1

     def count_non_empty_fields(self,dict):
         count = 0
         for value in dict.values():
             if value != None:  # 判断值是否为空
                 count += 1
         return count



class TypeTwoGetChunksNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    def process(self, chunks, start_index, next_start_index):
        result = []
        chunks_len = len(chunks)
        #如果沒有主要表格，直接將當前窗口内的excel算一個完整的块
        if chunks_len==0:
            result.append({
                "chunk": [start_index, next_start_index-1],
                "flag": True
            })
        else:
            for chunk in chunks:
                start = min(chunk[0],start_index)
                end = chunk[1]
                #如果這个块的最后一行不是当前窗口的最后一行，说明这个块肯定是完整的。
                if end < next_start_index-1:
                    result.append({
                        "chunk":[start,end],
                        "flag":True
                    })
                else:
                    # 如果這个块的最后一行是当前窗口的最后一行，说明这个块不一定是完整的。
                    result.append({
                        "chunk": [start, end],
                        "flag": False
                    })
                #下一个完整块的第一行肯定是当前完整快最后一行+1
                start_index= end+1
        return result

    async def __call__(self, state: State):
        llm_window = state["llm_window"]
        #当前从excel哪一行开始
        start_index =0
        next_start_index = 0
        stop = False
        chunks=[]
        excel_dict = state["excel_dict"]
        dict_len=len(state["excel_dict"])
        user_prompt=  f" Json 格式的 Excel 表格:"
        while stop==False:
            stop,message,start_index,next_start_index = get_type_two_window(excel_dict,llm_window,start_index,self.system_prompt,user_prompt)
            if stop:
                chunk = [start_index, next_start_index - 1]
                chunks.append(chunk)
                print(f"加载了{chunk}")
            else:
                try:
                    response = await OpenAIClientSingleton().create_completion(messages=message)
                    result = response.choices[0].message.content
                    result_json = extract_json(result)
                    #在指定模型上下文长度之内，主要excel块和额外excel块
                    main_chunks = [[int(x) for x in inner] for inner in result_json["main"]]
                    extra_chunks = [[int(x) for x in inner] for inner in result_json["extra"]]
                    #将两种块合并，得到一个个完整的块。对于当前的上下文窗口来说
                    temp_chunks=self.process(main_chunks,start_index,next_start_index)
                    #将所有的合并块插入到一个列表中，如果当前块是主要块，但是不确定是否被截断，则将游标左移。
                    for i,item in enumerate(temp_chunks):
                        chunk = item["chunk"]
                        flag = item["flag"]
                        #如果是完整的主要快或者全部是次要块，则插入。
                        if flag == True or len(temp_chunks)==1:
                            chunks.append(chunk)
                            print(f"加载了{chunk}")
                            # 如果顺利的将所有块插入了，则下次的工作窗口就是这次的最后一行的下一行。
                            if i == len(temp_chunks) - 1:
                                start_index = chunk[1] + 1
                                break
                        else:
                            # 如果不是完整的主要快或者全部是次要块，则将游标左移。
                            start_index = chunk[0]
                            break
                except Exception as e:
                    exception_message = str(e)
                    return Command(update={"error": f"TypeTwoGetChunksNode出错：{exception_message}"}, goto=END)
        return Command(update={"chunks": chunks}, goto="TypeTwoGetPromptsBot")


class TypeTwoGetPromptsNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt

    def __call__(self, state: State):
        llm_window = state["llm_window"]
        excel_dict = state["excel_dict"]
        chunks =state["chunks"]
        user_prompt = f"'json格式的excel文件:"
        result = get_parse_window(chunks,excel_dict,llm_window,self.system_prompt,user_prompt)
        return Command(update={"type_two_exe_prompts": result}, goto="TypeTwoExeBot")

class TypeTwoExeNode():
    async def __call__(self, state: State):
        json_list=[]
        type_two_exe_prompt=state["type_two_exe_prompts"]
        try:
            for message in type_two_exe_prompt:
                response =await OpenAIClientSingleton().create_completion(messages=message[1])
                result = response.choices[0].message.content
                result_json = extract_json_list(result)
                json_list.extend(result_json)


            return Command(update={"result": json_list}, goto=END)
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"TypeTwoExeNode出错：{exception_message}"}, goto=END)








class OnlyllmNode:
    def __init__(self, system_prompt):
        self.system_prompt = system_prompt
    async def __call__(self, state: State):

        human_input = f"请从我提供的 JSON 格式的 Excel 表格中，准确识别并提取所有合法的电费结算单记录，输出格式必须为一个 JSON 数组。JSON输入:{state["excel_dict"]}/no_think。"
        message = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": human_input}]
        try:
            response =await OpenAIClientSingleton().create_completion(messages=message)
            result = response.choices[0].message.content
            result_json = extract_json_list(result)

            return Command(update={"result":result_json},goto=END)
        except Exception as e:
            exception_message = str(e)
            return Command(update={"error": f"OnlyllmNode出错：{exception_message}"}, goto=END)



class BuildGraph:
    def __init__(self):
        self.graph = self.build_graph()

    def build_graph(self):
        config  = Config()
        graph_builder = StateGraph(State)
        PreProcessBot = PreProcessNode()

        FineTypeBot = FineTypeNode(config.type_system)
        TypeOneGetHeadBot = TypeOneGetHeadNode(config.get_head_system)
        TypeOneGetIndexBot = TypeOneGetIndexNode(config.get_index_system)
        TypeOneGetCommonBot = TypeOneGetCommonNode(config.get_head_common_system)
        TypeOneExeBot = TypeOneExeNode()



        TypeTwoGetChunksBot =TypeTwoGetChunksNode(config.depart_system)
        TypeTwoGetPromptsBot=TypeTwoGetPromptsNode(config.prompt)
        TypeTwoExeBot = TypeTwoExeNode()

        OnlyllmBot = OnlyllmNode(config.prompt)

        graph_builder.add_node("PreProcessBot", PreProcessBot)
        graph_builder.add_node("FineTypeBot", FineTypeBot)


        graph_builder.add_node("TypeOneGetHeadBot", TypeOneGetHeadBot)
        graph_builder.add_node("TypeOneGetIndexBot", TypeOneGetIndexBot)
        graph_builder.add_node("TypeOneGetCommonBot", TypeOneGetCommonBot)
        graph_builder.add_node("TypeOneExeBot", TypeOneExeBot)

        graph_builder.add_node("TypeTwoGetChunksBot",TypeTwoGetChunksBot)
        graph_builder.add_node("TypeTwoGetPromptsBot", TypeTwoGetPromptsBot)
        graph_builder.add_node("TypeTwoExeBot", TypeTwoExeBot)

        graph_builder.add_node("OnlyllmBot", OnlyllmBot)


        graph_builder.add_edge(START, "PreProcessBot")
        graph = graph_builder.compile()
        return graph

    def get_graph(self):
        return self.graph



if __name__ == "__main__":

    excel_data = pd.read_excel("../excel/最最特殊20250102171501269.xlsx", header=None, sheet_name=None)
    excel_df=  [df for sheet_name, df in excel_data.items()]
    df = excel_df[0]
    graph = BuildGraph().get_graph()
    async def get_item():
        result = await graph.ainvoke(input={
            "type_window": 11000,
            "llm_window": 6500,

            "excel_df":df,
            "excel_dict": None,
            "excel_list":None,
            "excel_type": None,
            "type_one_head":None,
            "type_one_index":None,


            'type_two_exe_prompts': None,
            'chunks':  None,
            'result':  None,
        })  # 调用异步函数并等待结果
        print(f"获取的结果是: {result}")
    #
    #
    # # 运行异步事件循环
    asyncio.run(get_item())  # 在 Python 3.7+ 中，使用 asyncio.run 来执行主函数

    # def blocking_task(graph,n,excel):
    #     print(f"Task {n} starting")
    #     result = graph.invoke(input={
    #         "excel_df": excel_df[0],
    #         "excel_json": None,
    #         "excel_type": None}, output_keys="result")
    #     print(f"Task {n} finished")
    #     return result
    #
    # async def compute():
    #     wb = load_workbook('税金.xlsx', read_only=True)
    #     excel_df=[]
    #     for sheet in wb.sheetnames:
    #         excel_df.append(pd.read_excel("./税金.xlsx",header=None,sheet_name=sheet))

    #     executor = ThreadPoolExecutor(max_workers=5)
    #     loop = asyncio.get_event_loop()
    #     return {"task_id": task_id, "result": result}
    #
    #     print(result)
    #
    #     tasks = [loop.run_in_executor(executor, blocking_task, graph,index,item) for index,item in enumerate(excel_df)]
    #     wb = load_workbook('最最特殊20250102171501269.xlsx', read_only=True)
    #     excel_df=[]
    #     for sheet in wb.sheetnames:
    #         excel_df.append(pd.read_excel("./最最特殊20250102171501269.xlsx",header=None,sheet_name=sheet))
    #     graph = BuildGraph().get_graph()
    #     # start = time.time()
    #     # result = graph.invoke(input={
    #     # "excel_df": excel_df[0],
    #     # "excel_json": None,
    #     # "excel_type": None}, output_keys="result")
    #     # end = time.time()
    #     # print(len(result))
    #     # print(end-start)
    #     for event in graph.stream(input={
    #         "excel_df":excel_df[0],
    #         "excel_json": None,
    #         "excel_type": None,
    #         "type_one_head":None,
    #         "type_one_index":None,
    #         'type_two_exe_prompt': None,
    #         'chunks':  None,
    #         'result':  None,
    #
    #     }):
    #         print("######################################")
    #         print("对话输出:", event)
    #         print('\n')


        #

    # compute()